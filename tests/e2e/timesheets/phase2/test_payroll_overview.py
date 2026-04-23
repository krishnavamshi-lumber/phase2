"""
Phase 2 — Payroll overview, preview, submit, and download reports.

Run this 4-5 hours after Phase 1, once the payroll overview is ready.

How to select which job to run:
  - Interactive (local): a numbered menu is shown at startup
  - Automated (Streamlit / CI): set env var PAYROLL_RUN_ID=<run_id>

Steps:
  8.  Authenticate + navigate to Payroll Overview
  8c. Wait for + verify employee names (up to 10 attempts)
  9.  Navigate to Payroll Preview and submit
  10. Download reports (Payroll Paycheck, Cash Requirement, Paystub)
"""

from __future__ import annotations

import json
import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Optional

import pytest
from playwright.sync_api import Page, expect

from fixtures.environments.environment_config import get_base_url
from utils.base.base_test import BaseTest
from utils.helpers.navigation_helper import NavigationHelper

from tests.e2e.timesheets.shared.helpers import (
    PAYROLL_TYPE,
    _normalize_date,
    select_pay_period_in_overview,
)
from tests.e2e.timesheets.shared.job_store import prompt_job_selection, mark_complete, mark_failed
# AFTER
import openpyxl
from tests.e2e.timesheets.shared.timesheet_loader import (
    load_timesheet_data,
    TimesheetUploadConfig,
    TimesheetUser,
)


# ── Upload button selector ────────────────────────────────────────────────────
# Set to 1 for old format (Emp No / First Name / Last Name columns in Excel)
# Set to 3 for new format (Employee Name column in timesheet_upload_3rdupload.xlsx)
UPLOAD_BUTTON = 3


# ── Button-3 data loader ──────────────────────────────────────────────────────
def _load_timesheet_data_button3(timesheet_file: str) -> TimesheetUploadConfig:
    _PROJECT_ROOT = Path(__file__).parents[4]
    file_path = str(
        _PROJECT_ROOT / "data-driven-files" / "timesheets" / timesheet_file
    )
    print(f"\n[INFO] Loading button-3 timesheet data from: {file_path}")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Timesheet file not found at: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    DATE_FMT = "%m-%d-%Y %I:%M %p"

    users: list[TimesheetUser] = []
    seen: set[str] = set()
    all_dates: list[datetime] = []

    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        raw = {h: v for h, v in zip(headers, excel_row)}

        emp_name = str(raw.get("Employee Name") or "").strip()
        if not emp_name:
            continue

        parts = emp_name.split(" ", 1)
        first_name = parts[0]
        last_name = parts[1] if len(parts) > 1 else ""

        if emp_name not in seen:
            seen.add(emp_name)
            users.append(
                TimesheetUser(emp_id="", first_name=first_name, middle_name="", last_name=last_name)
            )

        for col_name in ("Clocked In At", "Clocked Out At"):
            val = raw.get(col_name)
            if val:
                if isinstance(val, datetime):
                    all_dates.append(val)
                else:
                    try:
                        all_dates.append(datetime.strptime(str(val).strip(), DATE_FMT))
                    except ValueError:
                        pass

    if not users:
        raise ValueError("No employee data found in button-3 timesheet file")

    if not all_dates:
        raise ValueError("No valid dates found in Clocked In At / Clocked Out At columns")

    start_date = min(all_dates).strftime("%Y-%m-%d")
    end_date = max(all_dates).strftime("%Y-%m-%d")

    print(f"[INFO] Loaded {len(users)} unique user(s)")
    print(f"[INFO] Date range: {start_date} to {end_date}\n")

    return TimesheetUploadConfig(
        file_path=file_path,
        start_date=start_date,
        end_date=end_date,
        users=users,
        rows=[],
    )


# ── Structured run logger ─────────────────────────────────────────────────────
class _StepLog:
    """Timestamped, sectioned log for one Phase 2 run — saved alongside PDFs."""

    def __init__(self) -> None:
        self._lines: list[str] = []
        self._missing_employees: list[str] = []

    def _ts(self) -> str:
        return datetime.now().strftime("%H:%M:%S")

    def header(self, run_id: str, username: str, start: str, end: str) -> None:
        sep = "=" * 80
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for ln in [sep, "PHASE 2 PAYROLL RUN LOG", sep,
                   f"Run ID   : {run_id}", f"Username : {username}",
                   f"Period   : {start} to {end}", f"Started  : {now}", ""]:
            self._lines.append(ln)

    def section(self, title: str) -> None:
        sep = "=" * 80
        for ln in ["", sep, title, sep]:
            self._lines.append(ln)

    def ok(self, msg: str) -> None:
        self._lines.append(f"[{self._ts()}] [OK]   {msg}")

    def info(self, msg: str) -> None:
        self._lines.append(f"[{self._ts()}] [INFO] {msg}")

    def warn(self, msg: str) -> None:
        self._lines.append(f"[{self._ts()}] [WARN] {msg}")

    def record_missing(self, names: list[str]) -> None:
        self._missing_employees = names

    def get_missing(self) -> list[str]:
        return list(self._missing_employees)

    def save(self, path: str) -> None:
        sep = "=" * 80
        self._lines += ["", sep,
                        f"Log saved: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                        sep]
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(self._lines))
        print(f"[OK] Log saved: {path}")


# ── Select the job at module load time ────────────────────────────────────────
print("\nPhase 2 -- selecting payroll job...")
_SELECTED_JOB: dict = prompt_job_selection()
print(f"Phase 2 -- job selected: run_id={_SELECTED_JOB['run_id']}\n")

# ── Load timesheet data for employee verification ─────────────────────────────
print("Phase 2 -- loading timesheet data for employee name verification...")
if UPLOAD_BUTTON == 1:
    _TIMESHEET_CONFIG: TimesheetUploadConfig = load_timesheet_data()
else:
    _TIMESHEET_CONFIG: TimesheetUploadConfig = _load_timesheet_data_button3(
        _SELECTED_JOB["timesheet_file"]
    )
print("Phase 2 -- initialization complete\n")


# ============================================================================
# TEST
# ============================================================================

class TestPayrollOverview:

    @pytest.fixture(autouse=True)
    def setup(self, page: Page):
        self.navigation = NavigationHelper(page)

    def test_payroll_overview_and_download(self, page: Page):
        job              = _SELECTED_JOB
        timesheet_config = _TIMESHEET_CONFIG
        run_id           = job["run_id"]

        print(f"\n{'=' * 80}")
        print("PHASE 2 -- Payroll Overview, Preview & Download")
        print(f"  run_id    : {run_id}")
        print(f"  username  : {job['username']}")
        print(f"  period    : {job['pay_period_start']} to {job['pay_period_end']}")
        print(f"{'=' * 80}\n")

        try:
            self._run(page, job, timesheet_config, run_id)
            mark_complete(run_id)
        except Exception as exc:
            mark_failed(run_id, str(exc))
            raise

    # ------------------------------------------------------------------

    def _run(
        self,
        page: Page,
        job: dict,
        timesheet_config: TimesheetUploadConfig,
        run_id: str,
    ) -> None:

        log = _StepLog()
        log.header(run_id, job["username"], job["pay_period_start"], job["pay_period_end"])

        # Create download dir immediately so the log is always saved
        download_dir = os.environ.get("REPORT_DOWNLOAD_DIR") or str(
            Path(__file__).parents[4]
            / "data-driven-files"
            / "timesheets"
            / "downloads"
            / datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        )
        os.makedirs(download_dir, exist_ok=True)
        log.info(f"Download directory: {download_dir}")

        try:
            self._run_steps(page, job, timesheet_config, run_id, log, download_dir)
        finally:
            log_path = os.path.join(download_dir, "phase2_run.log")
            log.save(log_path)

    def _run_steps(
        self,
        page: Page,
        job: dict,
        timesheet_config: TimesheetUploadConfig,
        run_id: str,
        log: "_StepLog",
        download_dir: str,
    ) -> None:

        # --------------------------------------------------------------------
        # STEP 8: Authenticate + go to Payroll Overview
        # --------------------------------------------------------------------
        log.section("STEP 8 — Authentication & Navigation")
        print("\nStep 8: Authenticate and navigate to Payroll Overview")
        print("-" * 40)

        page.goto(get_base_url(), timeout=60000)
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(3000)

        BaseTest(page).v3_perform_authentication(job["username"], job["password"])
        log.ok("Authentication completed")
        print("[OK] Authentication completed")

        overview_url = get_base_url().rstrip("/") + "/payroll/overview"
        page.goto(overview_url)
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(3000)
        log.ok(f"Navigated to Payroll Overview: {overview_url}")
        print(f"[OK] Navigated to Payroll Overview: {overview_url}\n")

        # --------------------------------------------------------------------
        # STEP 8c: Select pay period + verify employee names
        #          Retry ONLY when pay period not ready (up to 10 x 4 min).
        #          Missing employees are recorded in the log — never causes a retry.
        # --------------------------------------------------------------------
        log.section("STEP 8c — Pay Period Selection & Employee Verification")
        print("\nStep 8c: Select Pay Period & Verify Employee Names")
        print("-" * 40)
        print(f"   Checking {len(timesheet_config.users)} user(s) -- up to 10 attempts x 4 min")

        _MAX_ATTEMPTS      = 10
        _WAIT_MS           = 240_000   # 4 minutes per retry (pay period only)
        _all_found         = False
        _missing_employees: list[str] = []
        matched_pay_period: Optional[tuple[str, str]] = None

        for _attempt in range(1, _MAX_ATTEMPTS + 1):
            log.info(f"Attempt {_attempt}/{_MAX_ATTEMPTS}: selecting pay period")
            print(f"\n   Attempt {_attempt}/{_MAX_ATTEMPTS}: waiting for pay period...")

            page.wait_for_load_state("networkidle")
            page.wait_for_timeout(3000)

            try:
                matched_pay_period = select_pay_period_in_overview(
                    page,
                    job["pay_period_start"],
                    job["pay_period_end"],
                    pay_type=PAYROLL_TYPE,
                )
                _pp_str = (
                    f"{matched_pay_period[0]} to {matched_pay_period[1]}"
                    if matched_pay_period else "no period matched"
                )
                log.ok(f"Pay period selected: {_pp_str}")
                print(f"   Matched pay period: {_pp_str}")
            except RuntimeError as _e:
                log.warn(f"Pay period not available yet: {_e}")
                print(f"   [WARN] Pay period not available yet: {_e}")
                if _attempt < _MAX_ATTEMPTS:
                    log.info(f"Waiting {_WAIT_MS // 60000} min then refreshing...")
                    print(f"   [INFO] Waiting {_WAIT_MS // 60000} min then refreshing...")
                    page.wait_for_timeout(_WAIT_MS)
                    page.reload()
                continue

            page.wait_for_timeout(2000)

            # Check ALL employees — never break early on a missing name.
            # Missing employees are logged and the run proceeds.
            _found_names:   list[str] = []
            _missing_names: list[str] = []

            for _idx, user in enumerate(timesheet_config.users):
                first_name     = user.first_name.strip()
                last_name      = user.last_name.strip()
                base_locator   = page.locator(f'div[data-testid="employee-name-{first_name}"]')
                full_name_text = f"{last_name}, {first_name}"
                count  = base_locator.count()
                locator = (
                    base_locator.filter(has_text=full_name_text)
                    if count > 1
                    else base_locator
                )
                _timeout = 60_000 if _idx == 0 else 5_000
                try:
                    locator.wait_for(state="visible", timeout=_timeout)
                    _found_names.append(full_name_text)
                    log.ok(f"Employee found: {full_name_text}")
                    print(f"   [OK] Found: {full_name_text}")
                except Exception:
                    _missing_names.append(full_name_text)
                    log.warn(f"Employee not found on overview: {full_name_text}")
                    print(f"   [WARN] Not found: {full_name_text}")

            _missing_employees = _missing_names
            _all_found         = len(_missing_names) == 0

            _emp_summary = (
                f"{len(_found_names)}/{len(timesheet_config.users)} found"
                + (f" | Missing: {', '.join(_missing_names)}" if _missing_names else "")
            )
            log.info(f"Employee check complete: {_emp_summary}")
            print(f"   [INFO] {_emp_summary}")

            # Pay period is confirmed — always proceed regardless of missing employees
            break

        log.record_missing(_missing_employees)

        if _all_found:
            log.ok("All employees verified on overview page")
            print("[OK] Employee name verification done\n")
        else:
            log.warn(
                f"Proceeding with {len(_missing_employees)} unverified employee(s): "
                f"{_missing_employees}"
            )
            print("[WARN] Some employees not confirmed -- continuing with available data\n")

        # --------------------------------------------------------------------
        # STEP 9: Payroll Preview + Submit
        # --------------------------------------------------------------------
        log.section("STEP 9 — Payroll Preview & Submit")
        print("\nStep 9: Navigate to Payroll Preview")
        print("-" * 40)

        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(10000)
        print("[OK] Payroll Overview page loaded")

        preview_button = page.get_by_test_id("payroll-overview-preview-button")
        preview_button.wait_for(state="visible", timeout=10000)
        preview_button.click()
        log.ok("Clicked Preview button")
        print("[OK] Clicked Preview button")

        start_payroll_button = page.get_by_text("Start Payroll")
        try:
            start_payroll_button.wait_for(state="visible", timeout=30_000)
            start_payroll_button.click()
            print("[OK] Clicked Start Payroll")
        except Exception:
            print("[INFO] Start Payroll button not shown (skipped)")

        page.wait_for_url(re.compile(r"/payroll/preview"), timeout=120000)
        log.ok("Navigated to Payroll Preview page")
        print("[OK] Navigated to Payroll Preview page")

        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(3000)

        preview_submit_button = page.get_by_test_id("preview-payroll-submit-button")
        expect(preview_submit_button).to_be_visible(timeout=30_000)
        preview_submit_button.click()
        log.ok("Submitted payroll preview")
        print("[OK] Submitted payroll preview")

        submit_payroll_buttons = page.locator('//button[normalize-space()="Submit Payroll"]')
        expect(submit_payroll_buttons).to_have_count(2, timeout=10000)
        submit_payroll_buttons.nth(1).click()
        log.ok("Confirmed Submit Payroll in modal")
        print("[OK] Confirmed Submit Payroll in modal")

        page.wait_for_timeout(5000)

        close_button_submit = page.locator('//button[normalize-space()="Close"]')
        close_button_submit.wait_for(state="visible", timeout=30000)
        close_button_submit.click()
        log.ok("Closed submission modal")
        print("[OK] Closed submission modal")

        # Navigate to payroll history
        payroll_menu = page.get_by_test_id("nav-payroll-menu")
        payroll_menu.wait_for(state="visible", timeout=10000)
        payroll_menu.click()

        payroll_history = page.get_by_test_id("nav-payroll-history")
        payroll_history.wait_for(state="visible", timeout=10000)
        payroll_history.click()

        page.wait_for_timeout(3000)
        print("[OK] Payroll history page loaded")

        # Select the pay period in history
        select_pay_period_button = page.get_by_role("button", name="Select the pay period")
        expect(select_pay_period_button).to_be_visible(timeout=120_000)
        select_pay_period_button.click()
        print("[OK] Clicked Select the pay period button")

        target_date  = _normalize_date(datetime.strptime(job["pay_period_start"], "%Y-%m-%d").date())
        found_period = False
        max_attempts = 10

        for attempt in range(1, max_attempts + 1):
            print(f"[INFO] Attempt {attempt}/{max_attempts}")
            menu_items = page.get_by_role("menuitem")
            count = menu_items.count()

            for i in range(count):
                item = menu_items.nth(i)
                text = item.text_content() or ""

                if f"({PAYROLL_TYPE})" not in text:
                    continue

                match = re.search(r"(\d{1,2} \w{3} \d{4}) - (\d{1,2} \w{3} \d{4})", text)
                if not match:
                    continue

                try:
                    start = _normalize_date(datetime.strptime(match.group(1), "%d %b %Y").date())
                    end   = _normalize_date(datetime.strptime(match.group(2), "%d %b %Y").date())
                except ValueError:
                    continue

                if start <= target_date <= end:
                    item.click()
                    print(f"[OK] Selected matching period: {text}")
                    found_period = True
                    break

            if found_period:
                break

            view_more_button = page.get_by_role("menuitem", name="View More")
            if view_more_button.count() > 0:
                before = menu_items.count()
                view_more_button.click()
                print("[INFO] Clicked View More")
                wait_attempts = 0
                while menu_items.count() <= before and wait_attempts < 20:
                    page.wait_for_timeout(500)
                    wait_attempts += 1
            else:
                print("[WARN] No more items to load")
                break

        if not found_period:
            raise RuntimeError(
                f"No matching {PAYROLL_TYPE} period found for {job['pay_period_start']}"
            )

        # --------------------------------------------------------------------
        # STEP 10: Download Reports
        # --------------------------------------------------------------------
        log.section("STEP 10 — Report Downloads")
        print("\nStep 10: Download Reports")
        print("-" * 40)

        page.wait_for_timeout(7000)

        print(f"[OK] Download directory: {download_dir}")

        if matched_pay_period:
            _pp_start, _pp_end = matched_pay_period
        else:
            _pp_start = job["pay_period_start"]
            _pp_end   = job["pay_period_end"]
        _period_suffix = f"_{_pp_start}_to_{_pp_end}"
        print(f"   Period suffix: {_period_suffix}")

        with open(os.path.join(download_dir, "period_suffix.txt"), "w", encoding="utf-8") as _psf:
            _psf.write(_period_suffix)

        download_button = page.locator("//button[normalize-space()='Download']")
        expect(download_button).to_be_visible(timeout=30_000)
        download_button.click()
        print("[OK] Clicked Download button")

        # --- Payroll Paycheck ---
        paycheck_item  = page.locator("//li[normalize-space()='Payroll Paycheck']")
        paycheck_buffer: Optional[bytes] = None

        for attempt in range(1, 10):
            expect(paycheck_item).to_be_visible(timeout=15_000)
            with page.expect_response(
                lambda r: "get-payroll-paychecks" in r.url, timeout=30_000
            ) as paycheck_info:
                paycheck_item.click()

            paycheck_response = paycheck_info.value
            if paycheck_response.status == 200:
                paycheck_buffer = paycheck_response.body()
                print(f"[OK] Payroll Paycheck ready on attempt {attempt}")
                break

            print(f"[INFO] Payroll Paycheck not ready ({paycheck_response.status}), retrying... ({attempt}/9)")
            page.wait_for_timeout(120_000)

        if paycheck_buffer is None:
            raise RuntimeError("Payroll Paycheck failed to download after all attempts")

        paycheck_path = os.path.join(download_dir, f"payroll_paycheck{_period_suffix}.pdf")
        with open(paycheck_path, "wb") as f:
            f.write(paycheck_buffer)
        log.ok(f"Payroll Paycheck downloaded: {os.path.basename(paycheck_path)}")
        print(f"[OK] Payroll Paycheck downloaded: {paycheck_path}")

        # --- Cash Requirement ---
        cash_item = page.locator("//li[normalize-space()='Cash Requirement']")
        expect(cash_item).to_be_visible(timeout=15_000)

        with page.expect_response(
            lambda r: "get-cash-requirement" in r.url and r.status == 200,
            timeout=240_000,
        ) as cash_info:
            cash_item.click()

        cash_buffer = cash_info.value.body()
        cash_path   = os.path.join(download_dir, f"cash_requirement{_period_suffix}.pdf")
        with open(cash_path, "wb") as f:
            f.write(cash_buffer)
        log.ok(f"Cash Requirement downloaded: {os.path.basename(cash_path)}")
        print(f"[OK] Cash Requirement downloaded: {cash_path}")
         

	# --- Paystub ---
        paystub_item = page.locator("//li[normalize-space()='Paystub']")
        paystub_buffer: Optional[bytes] = None

        for attempt in range(1, 10):
            expect(paystub_item).to_be_visible(timeout=15_000)
            with page.expect_response(
                lambda r: "get-payroll-paystubs" in r.url, timeout=30_000
            ) as paystub_info:
                paystub_item.click()

            paystub_response = paystub_info.value
            if paystub_response.status == 200:
                paystub_buffer = paystub_response.body()
                print(f"[OK] Paystub ready on attempt {attempt}")
                break

            print(f"[INFO] Paystub not ready ({paystub_response.status}), retrying... ({attempt}/9)")
            page.wait_for_timeout(240_000)

        if paystub_buffer is None:
            raise RuntimeError("Paystub failed to download after all attempts")

        paystub_path = os.path.join(download_dir, f"paystub{_period_suffix}.pdf")
        with open(paystub_path, "wb") as f:
            f.write(paystub_buffer)
        log.ok(f"Paystub downloaded: {os.path.basename(paystub_path)}")
        print(f"[OK] Paystub downloaded: {paystub_path}")

        log.ok("All reports downloaded successfully")
        print(f"\n[OK] All reports downloaded to: {download_dir}\n")

        # Save tag summary
        tag_counter      = Counter(row.tag for row in timesheet_config.rows if row.tag)
        tag_summary_path = os.path.join(download_dir, "tag_summary.json")
        with open(tag_summary_path, "w", encoding="utf-8") as _tf:
            json.dump(dict(tag_counter), _tf, indent=2)
        print(f"[OK] Tag summary saved: {tag_summary_path}")

        # --------------------------------------------------------------------
        # SUMMARY
        # --------------------------------------------------------------------
        log.section("SUMMARY")
        log.info(f"Status      : COMPLETE")
        log.info(f"run_id      : {run_id}")
        log.info(f"username    : {job['username']}")
        log.info(f"period      : {_pp_start} to {_pp_end}")
        log.info(f"downloads   : {download_dir}")
        _verified = len(timesheet_config.users) - len(_missing_employees)
        log.info(f"employees   : {_verified}/{len(timesheet_config.users)} verified on overview")
        if _missing_employees:
            log.warn(f"Missing employees on overview: {', '.join(_missing_employees)}")

        print(f"\n{'=' * 80}")
        print("PHASE 2 COMPLETE")
        print(f"{'=' * 80}")
        print(f"  run_id       : {run_id}")
        print(f"  username     : {job['username']}")
        print(f"  period       : {_pp_start} to {_pp_end}")
        print(f"  downloads    : {download_dir}")
        print(f"  total rows   : {len(timesheet_config.rows)}")
        print(f"  users        : {len(timesheet_config.users)}")
        for idx, u in enumerate(timesheet_config.users, start=1):
            _status = "(MISSING from overview)" if f"{u.last_name}, {u.first_name}" in _missing_employees else ""
            print(f"    [{idx}] {u.first_name} {u.last_name} (ID: {u.emp_id}) {_status}")
        print(f"{'=' * 80}\n")
