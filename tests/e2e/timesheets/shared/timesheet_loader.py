"""
Shared dataclasses and data-loading utilities.
Used by both Phase 1 and Phase 2.
"""

from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl


# ── Root of the project (4 levels up from this file) ──────────────────────────
# shared/ → timesheets/ → e2e/ → tests/ → root/
_PROJECT_ROOT = Path(__file__).parents[4]


# ============================================================================
# DATACLASSES
# ============================================================================

@dataclass
class TimesheetRow:
    emp_id: str
    first_name: str
    middle_name: str
    last_name: str
    date_worked: str
    earnings_code: str
    hours: str
    project: str
    project_description: str
    task: str
    task_description: str
    cost_code: str
    cost_type: str
    labor_item: str
    labor_item_description: str
    union: str
    card_type: str
    work_status: str
    shift: str
    special_pay: str
    sub_direct: str
    cert: str
    excl_flag: str
    flag2: str
    flag1: str
    end_date: str
    tag: str = ""


@dataclass
class TimesheetUser:
    emp_id: str
    first_name: str
    middle_name: str
    last_name: str


@dataclass
class TimesheetUploadConfig:
    file_path: str
    start_date: str
    end_date: str
    users: list[TimesheetUser]
    rows: list[TimesheetRow]


# ============================================================================
# LOADERS
# ============================================================================

def _str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def load_timesheet_data() -> TimesheetUploadConfig:
    file_path = str(
        _PROJECT_ROOT / "data-driven-files" / "timesheets" / "timesheet_upload.xlsx"
    )
    print(f"\n[INFO] Loading timesheet data from: {file_path}")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Timesheet Excel file not found at: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = [_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row_values: dict, *names: str) -> str:
        for name in names:
            if name in row_values:
                return _str(row_values[name])
        return ""

    rows: list[TimesheetRow] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        raw: dict[str, object] = {h: v for h, v in zip(headers, excel_row)}

        if not _str(raw.get("Emp ID") or raw.get("EmpID")):
            continue

        dw_raw = raw.get("Date Worked") or raw.get("DateWorked") or ""
        if isinstance(dw_raw, (datetime, date)):
            date_worked = dw_raw.strftime("%Y-%m-%d")
        else:
            date_worked = _str(dw_raw)

        ed_raw = raw.get("End Date") or raw.get("EndDate") or ""
        if isinstance(ed_raw, (datetime, date)):
            end_date_val = ed_raw.strftime("%Y-%m-%d")
        else:
            end_date_val = _str(ed_raw)

        tag_val = col(raw, "Tag") or "Untagged"

        rows.append(
            TimesheetRow(
                emp_id=col(raw, "Emp ID", "EmpID"),
                first_name=col(raw, "First Name", "FirstName"),
                middle_name=col(raw, "Middle Name", "MiddleName"),
                last_name=col(raw, "Last Name", "LastName"),
                date_worked=date_worked,
                earnings_code=col(raw, "Earnings Code", "EarningsCode"),
                hours=col(raw, "Hours"),
                project=col(raw, "Project"),
                project_description=col(raw, "Project Description", "ProjectDescription"),
                task=col(raw, "Task"),
                task_description=col(raw, "Task Description", "TaskDescription"),
                cost_code=col(raw, "Cost Code", "CostCode"),
                cost_type=col(raw, "Cost Type", "CostType"),
                labor_item=col(raw, "Labor Item", "LaborItem"),
                labor_item_description=col(raw, "Labor item Description", "LaborItemDescription"),
                union=col(raw, "Union"),
                card_type=col(raw, "Card Type", "CardType"),
                work_status=col(raw, "Work Status", "WorkStatus"),
                shift=col(raw, "Shift"),
                special_pay=col(raw, "Special Pay", "SpecialPay"),
                sub_direct=col(raw, "SubDirect"),
                cert=col(raw, "Cert"),
                excl_flag=col(raw, "ExclFlag"),
                flag2=col(raw, "Flag2"),
                flag1=col(raw, "Flag1"),
                end_date=end_date_val,
                tag=tag_val,
            )
        )

    if not rows:
        raise ValueError("Timesheet Excel file is empty or has no data rows")

    date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    valid_dates = sorted(r.date_worked for r in rows if r.date_worked and date_pattern.match(r.date_worked))

    if not valid_dates:
        raise ValueError('No valid dates found in "Date Worked" column')

    user_map: dict[str, TimesheetUser] = {}
    for row in rows:
        if row.emp_id and row.emp_id not in user_map:
            user_map[row.emp_id] = TimesheetUser(
                emp_id=row.emp_id,
                first_name=row.first_name,
                middle_name=row.middle_name,
                last_name=row.last_name,
            )

    config = TimesheetUploadConfig(
        file_path=file_path,
        start_date=valid_dates[0],
        end_date=valid_dates[-1],
        users=list(user_map.values()),
        rows=rows,
    )

    print(f"[INFO] Loaded {len(rows)} rows | {len(config.users)} unique user(s)")
    print(f"[INFO] Date range: {config.start_date} to {config.end_date}\n")
    return config


def load_credentials() -> dict[str, str]:
    cred_path = str(
        _PROJECT_ROOT / "data-driven-files" / "credentials" / "credentials.json"
    )
    print(f"[INFO] Loading credentials from: {cred_path}")

    with open(cred_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    creds = data.get("credentials", {})
    if not creds.get("username") or not creds.get("password"):
        raise ValueError("Invalid credentials structure — expected credentials.username and credentials.password")

    print("[INFO] Credentials loaded\n")
    return {"username": creds["username"], "password": creds["password"]}
